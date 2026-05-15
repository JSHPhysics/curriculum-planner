"""
Generate example_physics_spec.xlsx — a working import file matching the schema in SPEC.md §5.

Covers a representative slice of Edexcel 1PH0 Triple Physics: enough to exercise every
import path (multiple topics, sub-topics, multi-objective lessons, practicals, depth flags,
separate-only flags, empty optional cells).

Not exhaustive of the full spec — it is a development fixture, not a teaching document.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3A5F"
CREAM = "FBF7EE"
LINE = "D4CFC0"
WHITE = "FFFFFF"

wb = Workbook()
ws = wb.active
ws.title = "Spec"

# Column headers per SPEC.md §5.1
headers = [
    "Topic", "Lesson No.", "Lesson Title", "Sub-topic",
    "Objectives", "Practical", "Difficulty", "Extra-depth",
    "Separate science only?", "Paper", "Notes"
]

# Data — fields match SPEC.md §5.1. Objectives separated by newline within a cell.
# Lessons are unique by (Topic, Sub-topic, Lesson No.); multi-row lessons merge.
rows = [
    # ===== T1: Key concepts of physics =====
    ("Key concepts of physics", 1, "SI units and prefixes", "Units and measurement",
     "Use SI base units (m, kg, s, A) correctly\nConvert between prefixes from pico to giga\nWrite numbers in standard form to appropriate significant figures",
     "", 1, "", "", "1 & 2", "Foundation thread — revisit throughout."),

    ("Key concepts of physics", 2, "Scalars and vectors", "Units and measurement",
     "Distinguish between scalar and vector quantities\nGive examples of each (mass vs weight, distance vs displacement, speed vs velocity)\nRepresent vectors graphically using arrows",
     "", 2, "", "", "1 & 2", ""),

    # ===== T2: Motion and forces =====
    ("Motion and forces", 1, "Distance and displacement", "Kinematics",
     "Distinguish distance from displacement\nCalculate average speed from distance and time\nInterpret distance-time graphs for constant speed",
     "", 1, "", "", "1", ""),

    ("Motion and forces", 2, "Speed-time graphs", "Kinematics",
     "Plot speed-time graphs from data\nCalculate distance travelled as area under a speed-time graph\nInterpret motion described by a speed-time graph",
     "", 2, "", "", "1", ""),

    ("Motion and forces", 3, "Acceleration", "Acceleration and Newton's laws",
     "Define acceleration as rate of change of velocity\nApply v = u + at\nCalculate acceleration from gradient of velocity-time graph",
     "", 2, "", "", "1", ""),

    ("Motion and forces", 4, "Equations of motion", "Acceleration and Newton's laws",
     "Apply v² = u² + 2as to motion under constant acceleration\nSolve multi-step kinematics problems",
     "", 3, "", "", "1", ""),

    ("Motion and forces", 5, "Newton's first and second laws", "Acceleration and Newton's laws",
     "State Newton's first law\nApply F = ma to determine the resultant force, mass or acceleration\nIdentify when an object is in equilibrium",
     "CP1 Force and acceleration", 2, "", "", "1", ""),

    ("Motion and forces", 6, "Newton's third law and free body diagrams", "Acceleration and Newton's laws",
     "State Newton's third law\nIdentify Newton's third law pairs\nDraw free body diagrams for stationary and moving objects",
     "", 2, "", "", "1", ""),

    ("Motion and forces", 7, "Terminal velocity (depth)", "Acceleration and Newton's laws",
     "Explain terminal velocity in terms of changing resultant force\nApply Newton's laws to falling objects with air resistance\nSketch velocity-time graphs for falling objects",
     "", 3, "yes", "", "1", "Depth extension — not required for foundation tier."),

    # T2 sub-topic that's deliberately separate-science-only
    ("Motion and forces", 8, "Momentum and impulse", "Momentum",
     "Define momentum as p = mv\nApply conservation of momentum to collisions in one dimension\nCalculate impulse as change in momentum",
     "", 3, "", "yes", "1", "Triple-only content."),

    ("Motion and forces", 9, "Stopping distances", "Stopping distances",
     "Define thinking distance, braking distance and stopping distance\nIdentify factors affecting each\nEstimate stopping distances at common road speeds",
     "", 2, "", "", "1", ""),

    # ===== T3: Conservation of energy =====
    ("Conservation of energy", 1, "Energy stores and transfers", "Energy stores",
     "Identify energy stores (kinetic, gravitational, elastic, thermal, chemical, nuclear, electrostatic, magnetic)\nDescribe energy transfers between stores\nDraw energy transfer diagrams",
     "", 1, "", "", "1", ""),

    ("Conservation of energy", 2, "Kinetic and gravitational PE", "Energy calculations",
     "Calculate kinetic energy using KE = ½mv²\nCalculate gravitational PE using GPE = mgh\nSolve problems involving conversion between KE and GPE",
     "", 2, "", "", "1", ""),

    ("Conservation of energy", 3, "Specific heat capacity", "Energy calculations",
     "Define specific heat capacity\nCalculate thermal energy using E = mcΔθ\nMeasure specific heat capacity experimentally",
     "CP2 Specific heat capacity", 2, "", "", "1", ""),

    ("Conservation of energy", 4, "Conservation and efficiency", "Efficiency",
     "State the principle of conservation of energy\nCalculate efficiency as useful energy / total energy\nIdentify wasted energy in real systems",
     "", 2, "", "", "1", ""),

    # ===== T4: Waves =====
    ("Waves", 1, "Wave properties", "Wave basics",
     "Define amplitude, wavelength, frequency, period and wave speed\nDistinguish transverse and longitudinal waves\nGive examples of each type",
     "", 1, "", "", "1", ""),

    ("Waves", 2, "The wave equation", "Wave basics",
     "Apply v = f × λ to calculate wave speed, frequency or wavelength\nMeasure wavelength and frequency of waves experimentally",
     "CP3 Waves investigation", 2, "", "", "1", ""),

    ("Waves", 3, "Reflection and refraction", "Wave behaviour",
     "Apply the law of reflection (angle of incidence = angle of reflection)\nDescribe refraction qualitatively\nDraw ray diagrams for refraction at a plane boundary",
     "", 2, "", "", "1", ""),

    ("Waves", 4, "Snell's law (depth)", "Wave behaviour",
     "Apply Snell's law n₁ sin θ₁ = n₂ sin θ₂\nCalculate refractive index from wave speeds\nExplain total internal reflection",
     "", 3, "yes", "", "1", "Depth extension."),

    ("Waves", 5, "Sound waves", "Sound and ultrasound",
     "Describe sound as a longitudinal pressure wave\nState the range of human hearing\nDescribe applications of ultrasound in medicine and industry",
     "", 1, "", "yes", "1", "Triple-only sub-topic."),

    # ===== T6: Radioactivity =====
    ("Radioactivity", 1, "Atomic structure", "Atomic model",
     "Describe the nuclear model of the atom (Rutherford / Bohr)\nState relative charges and masses of protons, neutrons and electrons\nUse nuclear notation",
     "", 1, "", "", "1", ""),

    ("Radioactivity", 2, "Isotopes", "Atomic model",
     "Define isotope\nIdentify isotopes from nuclear notation\nState that isotopes have the same chemical properties but different physical properties",
     "", 1, "", "", "1", ""),

    ("Radioactivity", 3, "Types of radiation", "Radioactive decay",
     "Describe alpha, beta, gamma and neutron radiation\nState their relative penetrating powers and ionising abilities\nIdentify suitable absorbers for each",
     "", 2, "", "", "1", ""),

    ("Radioactivity", 4, "Half-life", "Radioactive decay",
     "Define half-life\nCalculate half-life from a decay curve\nUse half-life to predict remaining activity after a given time",
     "", 3, "", "", "1", ""),

    ("Radioactivity", 5, "Nuclear equations", "Radioactive decay",
     "Balance nuclear equations for alpha and beta decay\nIdentify the daughter nuclide from a given decay\nDistinguish nuclear fission from nuclear fusion",
     "", 3, "", "", "1", ""),
]

# Write headers
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(*[Side(border_style="thin", color=LINE)] * 4)

ws.row_dimensions[1].height = 32

# Write rows
for r_idx, row in enumerate(rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = Border(*[Side(border_style="thin", color=LINE)] * 4)
    # Row height proportional to the longest objectives cell
    obj_text = row[4] if len(row) > 4 else ""
    line_count = max(1, obj_text.count("\n") + 1)
    ws.row_dimensions[r_idx].height = max(24, line_count * 18)

# Column widths
widths = [22, 9, 28, 28, 56, 24, 9, 11, 16, 8, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Save
import os
os.makedirs("/home/claude/curriculum_planner/examples", exist_ok=True)
out = "/home/claude/curriculum_planner/examples/example_physics_spec.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {len(rows)}")
print(f"Topics: {len(set(r[0] for r in rows))}")
print(f"Sub-topics: {len(set((r[0], r[3]) for r in rows))}")
